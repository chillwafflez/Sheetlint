"""Shared test fixtures.

The async client uses httpx + ASGITransport per the FastAPI best-practices
guide. `app.router.lifespan_context` runs the lifespan manually because
ASGITransport does not propagate lifespan events on its own.
"""

from __future__ import annotations

import io
from collections.abc import AsyncIterator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook

from sheetlint.config import Environment, Settings, get_settings
from sheetlint.main import app


@pytest.fixture
def test_settings() -> Settings:
    """Test-isolated settings: no AI, slow cleanup so it doesn't fight tests."""
    return Settings(
        environment=Environment.TEST,
        anthropic_api_key=None,
        enable_ai=False,
        job_cleanup_interval_seconds=3600,
    )


@pytest_asyncio.fixture
async def client(test_settings: Settings) -> AsyncIterator[AsyncClient]:
    app.dependency_overrides[get_settings] = lambda: test_settings
    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
            yield ac
    app.dependency_overrides.clear()


@pytest.fixture
def broken_workbook_bytes() -> bytes:
    """Tiny workbook seeded with one finding per major detector category.

    Kept minimal so tests stay fast — the full demo sample lives in
    samples/generate_sample.py.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    # Row 1 is a title (forces mid-sheet header detection)
    ws["A1"] = "Quarterly Report"
    headers = ["policy_id", "state", "premium", "agent_email"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)

    # Clean rows
    for i in range(20):
        excel_row = 3 + i
        ws.cell(row=excel_row, column=1, value=f"POL-{1000 + i:05d}")
        ws.cell(row=excel_row, column=2, value="California")
        ws.cell(row=excel_row, column=3, value=1000.0 + i)
        ws.cell(row=excel_row, column=4, value=f"agent{i:03d}@example.com")

    # Type-purity violation — one stringified currency in a numeric column
    ws.cell(row=10, column=3, value="$1,500.00")

    # Fuzzy state typo
    ws.cell(row=15, column=2, value="Californa")

    # Bad email pattern
    ws.cell(row=18, column=4, value="not-an-email")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
