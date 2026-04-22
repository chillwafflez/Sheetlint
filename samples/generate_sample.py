"""Generate the broken insurance-submissions sample file used by the demo.

Every issue this seeds is intentional and corresponds to one detector. Keep them
in sync — when you add a new detector, add a deliberate trigger here so the
demo continues to showcase the full surface.

Run from the project root:
    python samples/generate_sample.py
"""

from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

random.seed(42)

OUTPUT = Path(__file__).parent / "broken_insurance_submissions.xlsx"


def _generate_submissions_sheet(ws):
    """Main data sheet — seeded with one of every kind of issue."""
    ws.title = "Submissions"

    # --- Pre-header title rows (triggers structural: mid-sheet header shift)
    ws["A1"] = "Q3 Insurance Submissions Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Prepared by: J. Smith"
    ws["A3"] = ""

    # --- Header row at row 4 (header detection should find this)
    headers = [
        "policy_id",
        "submission_date",
        "state",
        "premium_amount",
        "agent_email",
        "claim_count",
        "total_paid",
        "monthly_revenue",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True)

    # --- Data rows starting at row 5
    n_rows = 150
    states = ["California", "Texas", "New York", "Florida", "Illinois"]
    fuzzy_typos = {"California": "Californa", "Texas": "Texsas"}  # fuzzy-duplicates trigger

    base_date = datetime(2024, 1, 1)
    revenue_baseline = 50000.0

    for i in range(n_rows):
        excel_row = i + 5

        # policy_id — clean
        ws.cell(row=excel_row, column=1, value=f"POL-{1000 + i:05d}")

        # submission_date — daily
        ws.cell(row=excel_row, column=2, value=base_date + timedelta(days=i))

        # state — sprinkle in a few fuzzy typos (rows 30, 60, 90)
        state = states[i % len(states)]
        if i in (30, 60, 90) and state in fuzzy_typos:
            state = fuzzy_typos[state]
        ws.cell(row=excel_row, column=3, value=state)

        # premium_amount — currency-as-text mixed in (statistical type-drift trigger).
        # Most rows are floats; rows 50-60 are stringified currency.
        if 50 <= i <= 60:
            ws.cell(row=excel_row, column=4, value=f"${random.randint(800, 2400):,}.00")
        else:
            ws.cell(row=excel_row, column=4, value=round(random.uniform(800, 2400), 2))

        # agent_email — pattern coverage trigger: mostly emails, two malformed
        if i in (12, 77):
            ws.cell(row=excel_row, column=5, value="not-an-email")
        else:
            ws.cell(row=excel_row, column=5, value=f"agent{i:03d}@example.com")

        # claim_count — clean integers
        ws.cell(row=excel_row, column=6, value=random.randint(0, 5))

        # total_paid — formula column. One row produces a #DIV/0! error.
        if i == 100:
            ws.cell(row=excel_row, column=7, value="=D105/0")  # forces #DIV/0!
        else:
            ws.cell(row=excel_row, column=7, value=f"=F{excel_row}*D{excel_row}")

        # monthly_revenue — time-series with a deliberate spike at row 80
        # plus a STUMPY discord around row 130 (an unusual *pattern*, not just a spike)
        value = revenue_baseline + random.gauss(0, 2000)
        if i == 80:
            value = revenue_baseline * 4.5  # obvious z-score spike
        elif 125 <= i <= 135:
            # Inverted V pattern — different shape than the rest of the series
            phase = (i - 125) / 10
            value = revenue_baseline + 25000 * (1 - abs(phase - 0.5) * 2)
        ws.cell(row=excel_row, column=8, value=round(value, 2))

    # --- Inject a duplicate row (exact-duplicates trigger)
    src_row = 5
    dup_row = n_rows + 5
    for col in range(1, len(headers) + 1):
        ws.cell(row=dup_row, column=col, value=ws.cell(row=src_row, column=col).value)

    # --- Merged cells in the data region (structural trigger)
    ws.merge_cells("C20:C22")


def _generate_hidden_sheet(wb):
    """Hidden sheet — should be flagged by the structural detector."""
    ws = wb.create_sheet("InternalNotes")
    ws["A1"] = "Internal field — do not share externally"
    ws["A2"] = "Loss ratio target: 0.62"
    ws.sheet_state = "hidden"


def _generate_clean_sheet(wb):
    """A small clean sheet to demonstrate the tool doesn't cry wolf."""
    ws = wb.create_sheet("Reference")
    ws["A1"] = "code"
    ws["B1"] = "description"
    refs = [
        ("AUTO", "Automobile policy"),
        ("HOME", "Homeowner policy"),
        ("LIFE", "Life insurance policy"),
        ("UMBR", "Umbrella policy"),
    ]
    for i, (code, desc) in enumerate(refs, start=2):
        ws.cell(row=i, column=1, value=code)
        ws.cell(row=i, column=2, value=desc)


def main():
    wb = Workbook()
    _generate_submissions_sheet(wb.active)
    _generate_hidden_sheet(wb)
    _generate_clean_sheet(wb)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print("Issues seeded:")
    print("  - Pre-header title rows (header at row 4 instead of 1)")
    print("  - Merged cells C20:C22 in data region")
    print("  - 11 stringified currency values in premium_amount (rows 50-60)")
    print("  - 2 malformed emails in agent_email")
    print("  - 3 fuzzy state typos (Californa, Texsas)")
    print("  - 1 duplicate row")
    print("  - 1 #DIV/0! formula error in total_paid")
    print("  - Time-series spike in monthly_revenue at row 80")
    print("  - Time-series shape discord at rows 125-135 (STUMPY)")
    print("  - 1 hidden sheet ('InternalNotes')")


if __name__ == "__main__":
    main()
