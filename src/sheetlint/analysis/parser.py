"""Excel parser that preserves structural signals pandas throws away.

We load each sheet twice: once with openpyxl (to see merged cells, formulas,
error cells, hidden state) and once as a pandas DataFrame (for stats/ML).
Detectors get a SheetView that exposes both views plus a best-guess header
row in case the file has pre-header title rows.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO

import openpyxl
import pandas as pd
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Excel error literals pandas hides from us
EXCEL_ERRORS = {"#REF!", "#N/A", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!"}


@dataclass
class FormulaCell:
    row: int  # 1-indexed, matches Excel
    column: int  # 1-indexed
    formula: str
    value: object  # cached calculated value (may be an error string)


@dataclass
class SheetView:
    """Combined pandas + openpyxl view of a single worksheet.

    `df` uses the detected header row; `header_row` is 1-indexed and reflects
    where that header actually lived in the raw sheet. If `header_row > 1`,
    there were title/metadata rows above the table.
    """

    name: str
    df: pd.DataFrame
    raw_values: list[list[object]]  # full sheet including pre-header rows
    header_row: int  # 1-indexed position of detected header row
    is_hidden: bool
    merged_ranges: list[str]  # e.g. ["A1:C1", "B5:B7"]
    formula_cells: list[FormulaCell]
    error_cells: list[tuple[int, int, str]]  # (row, col, error_literal), 1-indexed
    openpyxl_sheet: Worksheet = field(repr=False)

    @property
    def n_rows(self) -> int:
        return len(self.df)

    @property
    def n_cols(self) -> int:
        return len(self.df.columns)


@dataclass
class ExcelDocument:
    """A parsed workbook. Detectors iterate `sheets` and work with SheetViews."""

    filename: str
    sheets: list[SheetView]
    hidden_sheet_names: list[str]
    workbook: Workbook = field(repr=False)

    def sheet(self, name: str) -> SheetView | None:
        for s in self.sheets:
            if s.name == name:
                return s
        return None


def _detect_header_row(values: list[list[object]], max_scan: int = 10) -> int:
    """Find the most plausible header row in the first `max_scan` rows.

    Heuristic: the row with the highest fraction of non-empty string cells
    AND no numeric cells, among the first max_scan non-empty rows. Falls back
    to row 1 if no obvious candidate.
    """
    best_row, best_score = 1, -1.0
    for i, row in enumerate(values[:max_scan], start=1):
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if not cells:
            continue
        str_fraction = sum(isinstance(c, str) for c in cells) / len(cells)
        numeric_penalty = sum(isinstance(c, (int, float)) for c in cells) / len(cells)
        score = str_fraction - numeric_penalty + (len(cells) / 100)
        if score > best_score:
            best_score, best_row = score, i
    return best_row


def _read_sheet(ws: Worksheet, workbook_with_formulas: Workbook) -> SheetView:
    # Pull raw values first so we can detect the header row ourselves
    raw_values: list[list[object]] = [list(row) for row in ws.iter_rows(values_only=True)]
    # Trim fully-empty trailing rows
    while raw_values and all(c is None for c in raw_values[-1]):
        raw_values.pop()

    header_row = _detect_header_row(raw_values) if raw_values else 1

    if raw_values and header_row <= len(raw_values):
        headers = [
            str(c) if c is not None else f"Column_{i + 1}"
            for i, c in enumerate(raw_values[header_row - 1])
        ]
        data_rows = raw_values[header_row:]
        df = pd.DataFrame(data_rows, columns=headers)
        # Drop rows that are entirely empty (artifacts of open-ended sheets)
        df = df.dropna(how="all").reset_index(drop=True)
    else:
        df = pd.DataFrame()

    # Merged cells
    merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]

    # Formulas + error cells (use the formula-preserving workbook view)
    ws_formulas = workbook_with_formulas[ws.title]
    formula_cells: list[FormulaCell] = []
    error_cells: list[tuple[int, int, str]] = []
    for row in ws_formulas.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                # This is a formula (data_only=False view). Pair with calculated value.
                calculated = ws.cell(row=cell.row, column=cell.column).value
                formula_cells.append(FormulaCell(cell.row, cell.column, val, calculated))
                if isinstance(calculated, str) and calculated in EXCEL_ERRORS:
                    error_cells.append((cell.row, cell.column, calculated))
            elif isinstance(val, str) and val in EXCEL_ERRORS:
                error_cells.append((cell.row, cell.column, val))

    return SheetView(
        name=ws.title,
        df=df,
        raw_values=raw_values,
        header_row=header_row,
        is_hidden=ws.sheet_state != "visible",
        merged_ranges=merged_ranges,
        formula_cells=formula_cells,
        error_cells=error_cells,
        openpyxl_sheet=ws,
    )


def parse_excel(source: str | Path | BinaryIO | bytes, filename: str | None = None) -> ExcelDocument:
    """Parse an .xlsx file into an ExcelDocument.

    Accepts a path, a file-like object (e.g. Streamlit's UploadedFile), or raw bytes.
    Two openpyxl loads happen: `data_only=True` for calculated values, and
    `data_only=False` to preserve formula strings — detectors need both.
    """
    if isinstance(source, (str, Path)):
        path = Path(source)
        filename = filename or path.name
        wb_values = openpyxl.load_workbook(path, data_only=True, read_only=False)
        wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
    else:
        if isinstance(source, bytes):
            buf = io.BytesIO(source)
        else:
            data = source.read()
            buf = io.BytesIO(data)
        wb_values = openpyxl.load_workbook(buf, data_only=True, read_only=False)
        buf.seek(0)
        wb_formulas = openpyxl.load_workbook(buf, data_only=False, read_only=False)
        filename = filename or "uploaded.xlsx"

    sheets: list[SheetView] = []
    hidden_names: list[str] = []
    for ws in wb_values.worksheets:
        sv = _read_sheet(ws, wb_formulas)
        if sv.is_hidden:
            hidden_names.append(sv.name)
        sheets.append(sv)

    return ExcelDocument(
        filename=filename,
        sheets=sheets,
        hidden_sheet_names=hidden_names,
        workbook=wb_values,
    )
